from openpyxl.styles import Font, PatternFill, Border, Side

# Fonts
FONT_ARIAL_12 = Font(name='Arial', size=12)
BOLD_FONT = Font(name='Arial', size=12, bold=True)

# Fill Colors
HEADER_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
FOOTER_FILL = PatternFill(start_color="AAAAAA", end_color="AAAAAA", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Borders
THIN_BORDER = Side(border_style="thin", color="000000")
HEADER_BORDER = Border(bottom=THIN_BORDER, left=THIN_BORDER, right=THIN_BORDER)
FOOTER_BORDER = Border(top=THIN_BORDER, left=THIN_BORDER, right=THIN_BORDER)
VERTICAL_BORDER = Border(left=THIN_BORDER, right=THIN_BORDER)
