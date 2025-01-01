import os
import openpyxl
import re

# Define a constant for the separator
SEPARATOR = "##############################"

def generate_list():
    # Automatically find the input file matching the pattern "Ficha_Leilão_#####.xlsx"
    input_pattern = re.compile(r"Ficha_Leilão_\d{5}\.xlsx")
    input_file = None

    for file in os.listdir():
        if input_pattern.match(file):
            input_file = file
            break

    if not input_file:
        print("No matching input file found.")
        return

    # Extract the number from the input filename
    file_number = input_file.split('_')[-1].split('.')[0]
    output_file = f"Lista_Leilão_{file_number}.txt"

    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Open the output file for writing
    with open(output_file, 'w', encoding='utf-8') as f:
        # Iterate over the rows, skipping the header (assume header is in the first row)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row - 1):
            col_a = row[0].value
            col_e = row[4].value
            col_f = row[5].value
            col_g = row[6].value
            col_h = row[7].value

            # Get the formatted values for F, G, and H
            col_f_formatted = f'{col_f:,.2f}' if isinstance(col_f, (int, float)) else col_f
            col_g_formatted = f'{col_g:,.2f}' if isinstance(col_g, (int, float)) else col_g
            col_h_formatted = f'{col_h:,.2f}' if isinstance(col_h, (int, float)) else col_h

            # Write the formatted text to the output file
            f.write(f"{SEPARATOR}\n\n")
            f.write(f"{col_a}\n")
            f.write("-----\n")
            f.write(f"O valor da sua arremação é: R$ {col_g_formatted}\n")
            f.write(f"O valor do frete por {col_e} é: R$ {col_f_formatted}\n")
            f.write(f"O total a pagar é: R$ {col_h_formatted}\n\n")

    print(f"List generated successfully: {output_file}")

# Example usage
generate_list()
