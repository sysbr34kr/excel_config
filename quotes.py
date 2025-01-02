import configparser
import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment
from styles import (
    FONT_ARIAL_12, BOLD_FONT, HEADER_FILL, FOOTER_FILL,
    LIGHT_GRAY_FILL, WHITE_FILL, HEADER_BORDER, FOOTER_BORDER, VERTICAL_BORDER
)

def process_quotes():
    current_directory = os.getcwd()

    input_file = next((file for file in os.listdir(current_directory) if file.startswith("Arrematantes_Leilao_") and file.endswith(".xls")), None)
    if not input_file:
        raise FileNotFoundError("No input file matching 'Arrematantes_Leilao_#####.xls' found in the current directory.")
    input_file = os.path.join(current_directory, input_file)
    config_file = os.path.join(current_directory, "valores.ini")
    if not os.path.exists(config_file):
        raise FileNotFoundError("Configuration file 'valores.ini' not found in the current directory.")
    file_number = os.path.splitext(os.path.basename(input_file))[0].split('_')[-1]
    output_file = f"Cotações_Leilão_{file_number}.xlsx"
    config = configparser.ConfigParser()
    config.read(config_file)
    
    im_values = {
        '0.3Kg': float(config['IM_Weights']['IM_0_3Kg']),
        '0.9Kg': float(config['IM_Weights']['IM_0_9Kg']),
        '2.0Kg': float(config['IM_Weights']['IM_2_0Kg'])
    }
    
    df = pd.read_html(input_file, header=1)[0]
    df = df.iloc[:, [1, 6]]
    df.columns = ['Nome', 'CEP']
    df['Nome'] = df['Nome'].str.upper()
    df['CEP'] = df['CEP'].astype(str).str.replace("-", "", regex=False).apply(lambda x: x.zfill(8))
    df = pd.DataFrame({
        'Nome': df['Nome'],
        'CEP': df['CEP'],
        'Modalidade': '',
        'Peso': '',
        'Alt.': '',
        'Lar.': '',
        'Com.': '',
        'Valor': ''
    })

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = FONT_ARIAL_12
            cell.alignment = Alignment(horizontal="center" if r_idx == 1 else "left" if c_idx == 1 and r_idx > 1 else "center")

    modalidade_options = ["RETIRA", "IM", "PAC Min.", "PAC", "2x PAC", "SEDEX", "OUTRO"]
    dv = DataValidation(type="list", formula1=f"\"{','.join(modalidade_options)}\"", showDropDown=False)
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    for row in range(2, ws.max_row):
        dv.add(ws.cell(row=row, column=3))

    column_widths = {
        'A': 71.43, # Nome
        'B': 12.14, # CEP
        'C': 13.57, # Modalidade
        'D': 8.57,  # Peso
        'E': 10,    # Altura
        'F': 10,    # Largura
        'G': 10,    # Comprimento
        'H': 12.14  # Valor
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in range(2, ws.max_row + 1):
        modalidade_cell = ws.cell(row=row, column=3)
        peso_cell = ws.cell(row=row, column=4)
        altura_cell = ws.cell(row=row, column=5)
        largura_cell = ws.cell(row=row, column=6)
        comprimento_cell = ws.cell(row=row, column=7)
        valor_cell = ws.cell(row=row, column=8)

        peso_cell.value = (
            f'=IF({modalidade_cell.coordinate}="PAC Min.", "1,0 Kg", ' 
            f'IF({modalidade_cell.coordinate}="RETIRA", "-", ""))'
        )
        altura_cell.value = (
            f'=IF({modalidade_cell.coordinate}="PAC Min.", "-", ' 
            f'IF({modalidade_cell.coordinate}="IM", "-", ' 
            f'IF({modalidade_cell.coordinate}="RETIRA", "-", "")))'
        )
        largura_cell.value = (
            f'=IF({modalidade_cell.coordinate}="PAC Min.", "-", ' 
            f'IF({modalidade_cell.coordinate}="IM", "-", ' 
            f'IF({modalidade_cell.coordinate}="RETIRA", "-", "")))'
        )
        comprimento_cell.value = (
            f'=IF({modalidade_cell.coordinate}="PAC Min.", "-", ' 
            f'IF({modalidade_cell.coordinate}="IM", "-", ' 
            f'IF({modalidade_cell.coordinate}="RETIRA", "-", "")))'
        )
        valor_cell.value = (
            f'=IF({modalidade_cell.coordinate}="RETIRA", "-", ' 
            f'IF(AND({modalidade_cell.coordinate}="IM", ' 
            f'ISNUMBER(MATCH(LEFT({peso_cell.coordinate}, FIND(" ", {peso_cell.coordinate})-1), ' 
            f'{{"0,3","0,9","2,0"}}, 0))), ' 
            f'VLOOKUP(LEFT({peso_cell.coordinate}, FIND(" ", {peso_cell.coordinate})-1), ' 
            f'{{"0,3",{im_values["0.3Kg"]};"0,9",{im_values["0.9Kg"]};"2,0",{im_values["2.0Kg"]}}}, 2, FALSE), ""))'
        )

        valor_cell.number_format = '"R$" #,##0.00'
        peso_cell.number_format = '#,0.0 "Kg"'
        altura_cell.number_format = '##0 "cm"'
        largura_cell.number_format = '##0 "cm"'
        comprimento_cell.number_format = '##0 "cm"'

    summary_row = ws.max_row
    ws.cell(row=summary_row, column=1).value = f"=COUNTA(A2:A{ws.max_row - 1})"
    ws.cell(row=summary_row, column=1).font = BOLD_FONT
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=summary_row, column=8).value = f"=COUNTBLANK(H2:H{ws.max_row - 1})"
    ws.cell(row=summary_row, column=8).font = BOLD_FONT
    ws.cell(row=summary_row, column=8).alignment = Alignment(horizontal="center")
    ws.cell(row=summary_row, column=8).number_format = 'General'
    for col in range(2, ws.max_column):
        if col != 8:
            ws.cell(row=summary_row, column=col).value = ""
            ws.cell(row=summary_row, column=col).alignment = Alignment(horizontal="center")
    
    ws.freeze_panes = ws['A2']
    for cell in ws[1]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row - 1), start=2):
        fill = LIGHT_GRAY_FILL if row_idx % 2 == 0 else WHITE_FILL
        for cell in row:
            cell.fill = fill
    for cell in ws[ws.max_row]:
        cell.font = BOLD_FONT
        cell.fill = FOOTER_FILL
    for cell in ws[1]:
        cell.border = HEADER_BORDER
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
        for cell in row:
            cell.border = VERTICAL_BORDER
    for cell in ws[ws.max_row]:
        cell.border = FOOTER_BORDER

    wb.save(output_file)
    print(f"File saved as {output_file}")

if __name__ == "__main__":
    process_quotes()
