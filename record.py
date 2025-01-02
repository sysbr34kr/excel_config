import configparser
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment
from styles import (
    FONT_ARIAL_12, BOLD_FONT, HEADER_FILL, FOOTER_FILL,
    LIGHT_GRAY_FILL, WHITE_FILL, HEADER_BORDER, FOOTER_BORDER, VERTICAL_BORDER
)

def process_record():
    current_directory = os.getcwd()
    arrematantes_file = next((file for file in os.listdir(current_directory) if file.startswith("Arrematantes_Leilao_") and file.endswith(".xls")), None)
    cotacoes_file = next((file for file in os.listdir(current_directory) if file.startswith("Cotações_Leilão_") and file.endswith(".xlsx")), None)
    if not arrematantes_file or not cotacoes_file:
        raise FileNotFoundError("Matching 'Arrematantes_Leilao_#####' and 'Cotações_Leilão_#####' files not found.")
    file_number = os.path.splitext(os.path.basename(arrematantes_file))[0].split('_')[-1]
    if file_number != os.path.splitext(os.path.basename(cotacoes_file))[0].split('_')[-1]:
        raise ValueError("File numbers for 'Arrematantes_Leilao' and 'Cotações_Leilão' do not match.")
    config_file = os.path.join(current_directory, "valores.ini")
    if not os.path.exists(config_file):
        raise FileNotFoundError("Configuration file 'valores.ini' not found in the current directory.")
    config = configparser.ConfigParser()
    config.read(config_file)
    pacote_extra = float(config['Pacote_Extra']['pacote_extra'])
    comissao = float(config['Comissao']['comissao'])
    seguro = float(config['Seguro']['seguro'])

    arrematantes_df = pd.read_html(arrematantes_file, header=1)[0]
    cotacoes_df = pd.read_excel(cotacoes_file)
    cotacoes_df = cotacoes_df.dropna(subset=['Nome', 'CEP', 'Modalidade', 'Valor']).reset_index(drop=True)
    arrematantes_df = arrematantes_df.iloc[:, [0, 5, 11]]
    arrematantes_df.columns = ['Cartela', 'UF', 'Arrematação']
    cotacoes_df['CEP'] = cotacoes_df['CEP'].astype(int).astype(str).apply(lambda x: x.zfill(8))
    cotacoes_df['Valor'] = pd.to_numeric(cotacoes_df['Valor'], errors='coerce')
    arrematantes_df['Arrematação'] = (
    arrematantes_df['Arrematação']
    .str.replace('.', '', regex=False)
    .str.replace(',', '', regex=False)
    .astype(float) / 100
    )
    arrematantes_df['Arrematação'] = pd.to_numeric(arrematantes_df['Arrematação'], errors='coerce')
    arrematantes_df['Arrematação'] = arrematantes_df['Arrematação'] * (1 + (comissao / 100))
     
    new_df = pd.DataFrame({
      'Nome': cotacoes_df['Nome'],
      'Cartela': arrematantes_df['Cartela'],
      'CEP': cotacoes_df['CEP'],
      'UF': arrematantes_df['UF'],
      'Modalidade': cotacoes_df['Modalidade'],
      'Valor Env.': cotacoes_df.apply(
        lambda row: '-' if row['Modalidade'] == 'RETIRA' else row['Valor'] + (pacote_extra if row['Modalidade'] == 'PAC' else 2 * pacote_extra if row['Modalidade'] == '2x PAC' else 0),
        axis=1
        ),
      'Arrematação': arrematantes_df['Arrematação'],
      'Total': '',
      'Situação': '',
      'Observação': ''
    })
    
    new_df['Total'] = new_df.apply(
        lambda row: row['Arrematação']
        if row['Valor Env.'] == '-'
        else row['Arrematação'] + row['Valor Env.'] + (row['Arrematação'] * (seguro / 100)),
        axis=1
    )
    
    output_file = f"Ficha_Leilão_{file_number}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha"

    header = [
        "Nome", "Cartela", "CEP", "UF",
        "Modalidade", "Valor Env.",
        "Arrematação", "Total",
        "Situação", "Observação"
    ]
    for col_num, header_text in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(new_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = FONT_ARIAL_12
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")
            if c_idx in [6, 7, 8]:
                if isinstance(value, (int, float)):
                    cell.number_format = '"R$" #,##0.00'
    
    situacao_options = ["PG Arrematação", "PG Arrem. + Env.", "PG Desistência", "Outro"]
    dv = DataValidation(type="list", formula1=f"\"{','.join(situacao_options)}\"", showDropDown=False)
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    for row in range(2, ws.max_row):
        dv.add(ws.cell(row=row, column=9))
    column_widths = {
        'A': 71.43,  # Nome
        'B': 9.28,   # Cartela
        'C': 12,     # CEP
        'D': 5,      # UF
        'E': 13.57,  # Modalidade
        'F': 15,     # Valor Env.
        'G': 16.50,  # Arrematação
        'H': 16.50,  # Total
        'I': 21.42,  # Situação
        'J': 86      # Observação
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    summary_row = ws.max_row
    ws.cell(row=summary_row, column=1).value = f"=COUNTA(A2:A{ws.max_row - 1})"
    ws.cell(row=summary_row, column=5).value = "Total"
    ws.cell(row=summary_row, column=6).value = f"=SUM(F2:F{ws.max_row - 1})"
    ws.cell(row=summary_row, column=7).value = f"=SUM(G2:G{ws.max_row - 1})"
    ws.cell(row=summary_row, column=8).value = f"=SUM(H2:H{ws.max_row - 1})"
    ws.cell(row=summary_row, column=9).value = (
    f'=COUNTIFS(I2:I{summary_row - 1}, "")'
    f'+ COUNTIFS(I2:I{summary_row - 1}, "PG Arrematação")'
    f'- COUNTIFS(F2:F{summary_row - 1}, "-", I2:I{summary_row - 1}, "PG Arrematação")'
    )
    for col in range(1, ws.max_column):
        if col > 1 and col <= 4 and col == 10:
            ws.cell(row=summary_row, column=col).value = ""
        ws.cell(row=summary_row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=summary_row, column=col).font = BOLD_FONT
    
    ws.freeze_panes = ws['A2']
    for cell in ws[1]:
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row - 1), start=2):
        fill = LIGHT_GRAY_FILL if row_idx % 2 == 0 else WHITE_FILL
        for cell in row:
            cell.fill = fill
    for cell in ws[ws.max_row]:
        cell.font = BOLD_FONT
        cell.fill = FOOTER_FILL
    for cell in ws[1]:
        cell.border = HEADER_BORDER
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
        for cell in row:
            cell.border = VERTICAL_BORDER
    for cell in ws[ws.max_row]:
        cell.border = FOOTER_BORDER

    # Sheet 2
    info_ws = wb.create_sheet(title="Info")
    info_ws["A1"] = "Comissão"
    info_ws["B1"] = f"=Ficha!G{summary_row}*{comissao/100}"
    info_ws["B3"] = "Envios"
    info_ws["C3"] = "Arrematação"
    info_ws["D3"] = "Comissão"
    info_ws["E3"] = "Total"
    info_ws["A4"] = "A Receber"
    info_ws["B4"] = (
    f"=Ficha!F{summary_row} - SUMIFS(Ficha!F2:F{summary_row-1}, Ficha!I2:I{summary_row-1}, \"PG Arrem. + Env.\") - SUMIFS(Ficha!F2:F{summary_row-1}, Ficha!I2:I{summary_row-1}, \"PG Desistência\")"
    )
    info_ws["C4"] = (
    f"=Ficha!G{summary_row} - SUMIFS(Ficha!G2:G{summary_row-1}, Ficha!I2:I{summary_row-1}, \"PG Arrematação\") - SUMIFS(Ficha!G2:G{summary_row-1}, Ficha!I2:I{summary_row-1}, \"PG Arrem. + Env.\") - SUMIFS(Ficha!G2:G{summary_row-1}, Ficha!I2:I{summary_row-1}, \"PG Desistência\")"
    )
    info_ws["D4"] = (
        f"=Info!B1 - SUMIFS(Ficha!G2:G{summary_row-1}, Ficha!I2:I{summary_row-1}, \"<>\" & \"\") * {comissao/100}"
    )
    info_ws["E4"] = (
        f"=Ficha!H{summary_row} - SUMIFS(Ficha!H2:H{summary_row-1}, Ficha!I2:I{summary_row-1}, \"<>\" & \"\")"
    )
    info_ws["A6"] = "RETIRA"
    info_ws["A7"] = "IM"
    info_ws["A8"] = "PAC Min."
    info_ws["A9"] = "PAC"
    info_ws["A10"] = "2x PAC"
    info_ws["A11"] = "SEDEX"
    info_ws["A12"] = "OUTRO"
    info_ws["B6"] = f"=COUNTIF(Ficha!E2:E{summary_row-1}, \"RETIRA\")"
    info_ws["B7"] = f"=COUNTIF(Ficha!E2:E{summary_row-1}, \"IM\")"
    info_ws["B8"] = f"=COUNTIF(Ficha!E2:E{summary_row-1}, \"PAC Min.\")"
    info_ws["B9"] = f"=COUNTIF(Ficha!E2:E{summary_row-1}, \"PAC\")"
    info_ws["B10"] = f"=COUNTIF(Ficha!E2:E{summary_row-1}, \"2x PAC\")"
    info_ws["B11"] = f"=COUNTIF(Ficha!E2:E{summary_row-1}, \"SEDEX\")"
    info_ws["B12"] = f"=COUNTIF(Ficha!E2:E{summary_row-1}, \"OUTRO\")"

    info_column_widths = {
    "A": 12.85,  # Comissão, A Receber
    "B": 16.50,  # Envios
    "C": 16.50,  # Arrematação
    "D": 16.50,  # Comissão
    "E": 16.50,  # Total
    }
    for col, width in info_column_widths.items():
        info_ws.column_dimensions[col].width = width

    for row in info_ws.iter_rows():
        for cell in row:
            cell.font = FONT_ARIAL_12
    info_ws["A1"].alignment = Alignment(horizontal="left")
    for row in range(1, 13):
        for col in range(2, 6):
            info_ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
    for col in range(1, 6):
        info_ws.cell(row=3, column=col).font = BOLD_FONT
        info_ws.cell(row=3, column=col).fill = HEADER_FILL
    for row in range(1, 13):
        info_ws[f"A{row}"].font = BOLD_FONT
        info_ws[f"A{row}"].fill = HEADER_FILL
    info_ws["B1"].number_format = '"R$" #,##0.00'
    info_ws["D4"].number_format = '"R$" #,##0.00'

    vertical_border_ranges = [
    ("A1", "B1"),
    ("A3", "E3"),
    ("A4", "E4"),
    ("A6", "B6"),
    ("A7", "B7"),
    ("A8", "B8"),
    ("A9", "B9"),
    ("A10", "B10"),
    ("A11", "B11"),
    ("A12", "B12"),
    ]

    for start, end in vertical_border_ranges:
        start_cell = info_ws[start]
        end_cell = info_ws[end]
        for row in range(start_cell.row, end_cell.row + 1):
            for col in range(start_cell.column, end_cell.column + 1):
                cell = info_ws.cell(row=row, column=col)
                cell.border = VERTICAL_BORDER
    
    info_ws["A2"].fill = WHITE_FILL
    info_ws["A3"].fill = WHITE_FILL
    info_ws["A5"].fill = WHITE_FILL
    info_ws["B1"].fill = LIGHT_GRAY_FILL
    for col in range(2, 6):
        info_ws.cell(row=4, column=col).fill = LIGHT_GRAY_FILL
    for row in range(6, 13):
        info_ws[f"B{row}"].fill = LIGHT_GRAY_FILL

    wb.save(output_file)
    print(f"File saved as {output_file}")

if __name__ == "__main__":
    process_record()
