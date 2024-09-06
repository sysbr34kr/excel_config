# Required imports
import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Find all .xls files in the current directory
html_files = glob.glob('*.xls')

if html_files:
    # Use the most recent .xls file
    latest_html_file = max(html_files, key=lambda x: os.path.getctime(x))

    # Read the HTML file into a DataFrame
    df = pd.read_html(latest_html_file, header=None)[0]  # [0] gets the first table, if there are multiple
    
    # Manually assign the correct headers
    df.columns = [
        "Cartela",
        "Nome",
        "Endereço",
        "Bairro",
        "Cidade/UF",
        "UF",
        "CEP",
        "Telefone",
        "Email",
        "CPF/CNPJ",
        "Valor",
        "Other1",
        "Other2",
        "Total"
    ]
    
    # Save it as .xlsx
    temp_xlsx_file = 'convert_file.xlsx'
    df.to_excel(temp_xlsx_file, index=False)
    print(f"Converted {latest_html_file} to {temp_xlsx_file}")

    # Treat the converted file as the latest .xlsx file
    latest_file = temp_xlsx_file

else:
    # Find all .xlsx files in the current directory
    xlsx_files = glob.glob('*.xlsx')

    if not xlsx_files:
        print("No .xls or .xlsx files found in the current directory.")
        exit()  # Exit the script if no files are found

    # Use the most recent .xlsx file
    latest_file = max(xlsx_files, key=lambda x: os.path.getctime(x))

# Read .xlsx file
df = pd.read_excel(latest_file, engine='openpyxl')
print(f"Imported {latest_file}")

# Columns to keep
columns_to_keep = [df.columns[1], df.columns[0], df.columns[6], df.columns[5], df.columns[11]]
df = df[columns_to_keep]

# Rename columns
df.columns = ["Nome", "Cartela", "CEP", "UF", "Arrematação"]

# Invert formatting decimal and separators
df["Arrematação"] = df["Arrematação"].str.replace(".", "", regex=False)
df["Arrematação"] = df["Arrematação"].str.replace(",", "", regex=False)

# Convert "Arrematação" to numeric and handle errors
df["Arrematação"] = pd.to_numeric(df["Arrematação"], errors='coerce') / 100

# Comissão 10%
df["Arrematação"] = df["Arrematação"] * 1.10

# Format "Arrematação" to show two decimal places
df["Arrematação"] = df["Arrematação"].round(2)

# Convert 'Nome' to uppercase
df['Nome'] = df['Nome'].str.upper()

# Process CEP values
df["CEP"] = df["CEP"].astype(str)  # Converts content to string
df["CEP"] = df["CEP"].str.replace("-", "")  # Remove hyphens
df["CEP"] = df["CEP"].apply(lambda x: x.zfill(8))  # Ensures 8 characters

# Add new columns
new_column_names = [
    "Valor Env.",
    "Modalidade",
    "Seguro",
    "Total S/S",
    "Total C/S",
    "Situação",
    "Observação"
]
for name in new_column_names:
    df[name] = ""

# Convert DataFrame to Excel (xlsx format)
df.to_excel('output_file.xlsx', index=False)

# Load the newly created .xlsx file to apply formatting
wb = load_workbook('output_file.xlsx')
ws = wb.active

# Set number format for currency rows
for row in range(2, ws.max_row + 1):
    for col_letter in ['E', 'F', 'H', 'I', 'J']:
        cell = ws[f'{col_letter}{row}']
        cell.number_format = 'R$ #,##0.00' # Ensure decimal places with prefix

# Determine the size of the DataFrame (rows and columns)
total_rows = ws.max_row
total_columns = ws.max_column

# Add the summary row
summary_row = {
    "Nome": f'=COUNTA(A2:A{total_rows})',           # Total of names
    "Cartela": "",                                  # No info
    "UF": "",                                       # No info
    "CEP": "",                                      # No info
    "Arrematação": f'=SUM(E2:E{total_rows})',       # Total sum
    "Valor Env.": f'=SUM(F2:F{total_rows})',        # Total sum
    "Modalidade": "",                               # No info
    "Seguro": f'=SUM(H2:H{total_rows})',            # Total sum
    "Total S/S": f'=SUM(I2:I{total_rows})',         # Total sum
    "Total C/S": f'=SUM(J2:J{total_rows})',         # Total sum
    "Situação": f'=COUNTIF(K2:K{total_rows}, "")',  # Count empty cells
}

# Append summary row to the worksheet
for col, value in enumerate(summary_row.values(), start=1):
    ws.cell(row=total_rows + 1, column=col).value = value

# Update total_rows after appending the summary row
total_rows = ws.max_row

# Center summary row
for col in range(1, total_columns + 1):
    ws.cell(row=total_rows, column=col).alignment = Alignment(horizontal='center')

# Apply font changes to all cells
font = Font(name='Arial', size=12)
for row in ws.iter_rows(min_row=1, max_row=total_rows, min_col=1, max_col=total_columns):
    for cell in row:
        cell.font = font

# Set center alignment for all columns except the first
for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
    for cell in ws[col_letter]:
        cell.alignment = Alignment(horizontal="center")

# Set column widths
column_widths = {
    'A': 71.43,  # Nome
    'B': 8.57,   # Cartela
    'C': 12.14,  # CEP
    'D': 5,      # UF
    'E': 14.28,  # Arrematação
    'F': 12.14,  # Valor Env.
    'G': 12.86,  # Modalidade
    'H': 10.71,  # Seguro
    'I': 14.28,  # Total S/S
    'J': 14.28,  # Total C/S
    'K': 28.57,  # Situação
    'L': 87.86   # Observação
}
for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze the first and summary rows
ws.freeze_panes = ws['A2']

# Define styles for header, summary and even rows
header_fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
header_font = Font(name='Arial', size=12, color='000000')

summary_fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
summary_font = Font(name='Arial', size=12, color='000000')

even_row_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

# Define border style
border_style = Side(border_style='thin', color='000000')
border = Border(left=border_style, right=border_style)

# Apply header formatting
for col in range(1, total_columns + 1):
    header_cell = ws.cell(row=1, column=col)
    header_cell.fill = header_fill
    header_cell.font = header_font
    header_cell.border = border

# Apply summary formatting
for col in range(1, total_columns + 1):
    summary_cell = ws.cell(row=total_rows, column=col)
    summary_cell.fill = summary_fill
    summary_cell.font = summary_font
    summary_cell.border = border

# Apply light gray to fill even rows and add borders to all rows
for row in range(1, total_rows + 1):
    for col in range(1, total_columns + 1):
        cell = ws.cell(row=row, column=col)
        if row % 2 == 0:
            cell.fill = even_row_fill  
        cell.border = border

# Add dynamic formulas to columns H, I and J
for row in range(2, total_rows):
    # Column H: (Column E + Column F) * 2%, if F is empty, shows 0
    ws[f'H{row}'] = f'=IF(F{row}="", 0, (E{row} + F{row}) * 0.02)'

    # Column I: Column E + Column F
    ws[f'I{row}'] = f'=E{row} + F{row}'

    # Column J: Column E + Column F + Column H:
    ws[f'J{row}'] = f'=E{row} + F{row} + H{row}'

# Save the updated workbook
wb.save('output_file.xlsx')